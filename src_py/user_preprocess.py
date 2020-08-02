import openpyxl
wb = openpyxl.load_workbook("user.xlsx")
wb2 = openpyxl.Workbook()
# Workbook()
ws = wb.active
ws2 = wb2.create_sheet("new", 0)
i = 0
for row in ws.rows:
    if (i == 0):
        print(i / 1000)
        i += 1
        ws2.append([cell.value for cell in row])
        continue
    if i % 1000 == 0:
        print(i / 1000)
    i += 1
    if int(row[2].value) >= int(row[3].value) or int(row[9].value) == 0:
        continue
    # ws2.append(row)
    ws2.append([cell.value for cell in row])
wb2.save("new.xlsx")

